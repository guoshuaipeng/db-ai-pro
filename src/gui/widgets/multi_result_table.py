"""
多结果表格组件（支持Tab切换）
"""
from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QDialog,
    QTextEdit,
    QDialogButtonBox,
    QHBoxLayout,
    QFileDialog,
    QMenu,
    QApplication,
    QLineEdit,
    QSpinBox,
    QGraphicsOpacityEffect,
)
from src.utils.toast import show_toast
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QFont, QTextCharFormat, QColor, QSyntaxHighlighter, QClipboard, QMouseEvent
from PyQt6.QtCore import QRegularExpression
from typing import List, Dict, Optional
import json
import re
import csv
import logging
from datetime import datetime, date, time
from decimal import Decimal
from pathlib import Path

logger = logging.getLogger(__name__)


class JSONHighlighter(QSyntaxHighlighter):
    """JSON语法高亮"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # 定义高亮规则
        self.highlighting_rules = []
        
        # 关键字（true, false, null）
        keyword_format = QTextCharFormat()
        keyword_format.setForeground(QColor(86, 156, 214))  # 蓝色
        keyword_format.setFontWeight(700)  # 粗体
        keywords = ["true", "false", "null"]
        for keyword in keywords:
            pattern = QRegularExpression(f"\\b{keyword}\\b")
            self.highlighting_rules.append((pattern, keyword_format))
        
        # 字符串（用引号包围的内容）
        string_format = QTextCharFormat()
        string_format.setForeground(QColor(206, 145, 120))  # 橙色
        pattern = QRegularExpression('"[^"\\\\]*(\\\\.[^"\\\\]*)*"')
        self.highlighting_rules.append((pattern, string_format))
        
        # 数字
        number_format = QTextCharFormat()
        number_format.setForeground(QColor(181, 206, 168))  # 绿色
        pattern = QRegularExpression("\\b\\d+(\\.\\d+)?\\b")
        self.highlighting_rules.append((pattern, number_format))
    
    def highlightBlock(self, text: str):
        """高亮文本块"""
        for pattern, format in self.highlighting_rules:
            iterator = pattern.globalMatch(text)
            while iterator.hasNext():
                match = iterator.next()
                self.setFormat(match.capturedStart(), match.capturedLength(), format)


class SingleResultTable(QWidget):
    """单个查询结果表格"""
    
    def __init__(self, parent=None, main_window=None, sql: str = None):
        super().__init__(parent)
        self.main_window = main_window  # 主窗口引用，用于执行SQL
        self.original_sql = sql  # 原始SQL查询（不带LIMIT）
        self.original_data: List[Dict] = []  # 原始数据（用于生成WHERE条件）
        self.execute_query_func = None  # 自定义的执行查询函数（用于新标签）
        self.auto_limit_added = False  # 标记是否自动添加了LIMIT
        
        # 分页相关
        self.all_data = []  # 存储当前页数据
        self.current_page = 1  # 当前页码
        self.page_size = 50  # 每页显示的行数
        self.total_pages = 1  # 总页数
        self.total_rows = 0  # 总行数
        self.server_side_paging = False  # 是否使用服务器端分页
        
        self.init_ui()
    
    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # 工具栏（导出按钮）- 放在显示区域上面的右边
        toolbar_layout = QHBoxLayout()
        
        toolbar_layout.addStretch()
        
        # 导出按钮（放在右边）
        self.export_btn = QPushButton("导出")
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.show_export_menu)
        toolbar_layout.addWidget(self.export_btn)
        
        layout.addLayout(toolbar_layout)
        
        # 结果表格
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        # 支持单元格选择和行选择
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        # 启用多选功能
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        # 启用编辑功能：双击或按F2编辑
        self.table.setEditTriggers(
            QTableWidget.EditTrigger.DoubleClicked | 
            QTableWidget.EditTrigger.SelectedClicked |
            QTableWidget.EditTrigger.EditKeyPressed
        )
        
        # 设置选中样式：当前行浅色，当前单元格深色
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;  /* 白色背景 */
                selection-background-color: #E3F2FD;  /* 选中行浅蓝色背景 */
                selection-color: #000000;  /* 选中文本颜色 */
                alternate-background-color: #F5F5F5;  /* 交替行背景色 */
            }
            QTableWidget::item {
                background-color: transparent;  /* 默认透明背景 */
            }
            QTableWidget::item:selected {
                background-color: #BBDEFB;  /* 选中单元格深蓝色背景 */
            }
            QTableWidget::item:focus {
                background-color: #90CAF9;  /* 当前焦点单元格更深的蓝色 */
                border: 1px solid #2196F3;  /* 蓝色边框 */
            }
        """)
        
        # 设置表头
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)
        # 连接表头点击事件，点击列名时复制列名
        header.sectionClicked.connect(self.on_header_clicked)
        
        # 列的最大宽度（像素）
        self.max_column_width = 400
        
        # 连接双击事件（用于编辑）
        self.table.itemDoubleClicked.connect(self.on_row_double_clicked)
        
        # 连接单元格编辑完成事件
        self.table.itemChanged.connect(self.on_item_changed)
        
        # 启用右键菜单
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        layout.addWidget(self.table)
        
        # 分页控件
        self.pagination_widget = self._create_pagination_widget()
        self.pagination_widget.setVisible(False)  # 默认隐藏
        layout.addWidget(self.pagination_widget)
        
        # 状态标签（已隐藏，状态信息显示到主窗口状态栏）
        self.status_label = QLabel("等待查询结果...")
        self.status_label.setStyleSheet("color: #666; padding: 5px; border-top: 1px solid #ddd;")
        self.status_label.hide()  # 隐藏状态标签
        
        # 保存原始数据（用于显示JSON）
        self.raw_data: List[Dict] = []
        
        # 标记是否正在更新数据（避免itemChanged触发）
        self._updating_data = False
        
        # 保存修改的单元格信息：{(row, col): (old_value, new_value)}
        self.modified_cells: Dict[tuple, tuple] = {}
        
        # 保存正在执行的UPDATE worker
        self.update_worker = None
        
        # 保存正在执行的分页查询 worker
        self.pagination_worker = None
        
        # 保存连接信息（用于分页查询）
        self.connection_string = None
        self.connect_args = None
        
        # 记录上次的列数（用于判断是否需要调整列宽）
        self._last_column_count = 0
    
    def _restore_table_opacity(self):
        """恢复表格透明度"""
        # 移除透明效果，恢复到完全不透明
        self.table.setGraphicsEffect(None)
    
    def _create_pagination_widget(self):
        """创建分页控件"""
        widget = QWidget()
        layout = QHBoxLayout()
        layout.setContentsMargins(5, 5, 5, 5)
        widget.setLayout(layout)
        
        # 信息标签（显示当前页/总页数，以及行数范围）
        self.page_info_label = QLabel("第 1/1 页 (显示 0-0 行，共 0 行)")
        self.page_info_label.setStyleSheet("color: #666; font-size: 12px;")
        layout.addWidget(self.page_info_label)
        
        layout.addStretch()
        
        # 每页显示行数
        layout.addWidget(QLabel("每页显示:"))
        self.page_size_spin = QSpinBox()
        self.page_size_spin.setRange(10, 1000)
        self.page_size_spin.setSingleStep(10)
        self.page_size_spin.setValue(self.page_size)
        self.page_size_spin.setFixedWidth(80)
        self.page_size_spin.setToolTip("设置每页显示的行数")
        self.page_size_spin.valueChanged.connect(self._on_page_size_changed)
        layout.addWidget(self.page_size_spin)
        
        layout.addWidget(QLabel(" 行  "))
        
        # 首页按钮
        self.first_page_btn = QPushButton("首页")
        self.first_page_btn.setFixedWidth(60)
        self.first_page_btn.clicked.connect(self._go_first_page)
        layout.addWidget(self.first_page_btn)
        
        # 上一页按钮
        self.prev_page_btn = QPushButton("上一页")
        self.prev_page_btn.setFixedWidth(70)
        self.prev_page_btn.clicked.connect(self._go_prev_page)
        layout.addWidget(self.prev_page_btn)
        
        # 页码输入
        layout.addWidget(QLabel("第"))
        self.page_input = QLineEdit()
        self.page_input.setFixedWidth(50)
        self.page_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.page_input.setText("1")
        self.page_input.setToolTip("输入页码后按回车跳转")
        self.page_input.returnPressed.connect(self._on_page_input)
        layout.addWidget(self.page_input)
        
        self.total_pages_label = QLabel("/ 1 页")
        layout.addWidget(self.total_pages_label)
        
        # 下一页按钮
        self.next_page_btn = QPushButton("下一页")
        self.next_page_btn.setFixedWidth(70)
        self.next_page_btn.clicked.connect(self._go_next_page)
        layout.addWidget(self.next_page_btn)
        
        # 末页按钮
        self.last_page_btn = QPushButton("末页")
        self.last_page_btn.setFixedWidth(60)
        self.last_page_btn.clicked.connect(self._go_last_page)
        layout.addWidget(self.last_page_btn)
        
        return widget
    
    def _update_pagination_controls(self):
        """更新分页控件状态"""
        # 更新按钮状态
        self.first_page_btn.setEnabled(self.current_page > 1)
        self.prev_page_btn.setEnabled(self.current_page > 1)
        self.next_page_btn.setEnabled(self.current_page < self.total_pages)
        self.last_page_btn.setEnabled(self.current_page < self.total_pages)
        
        # 更新页码显示
        self.page_input.setText(str(self.current_page))
        self.total_pages_label.setText(f"/ {self.total_pages} 页")
        
        # 更新信息标签
        if self.server_side_paging:
            # 服务器端分页：使用 total_rows
            if self.total_rows == 0:
                self.page_info_label.setText("第 0/0 页 (显示 0-0 行，共 0 行)")
            else:
                start_row = (self.current_page - 1) * self.page_size + 1
                end_row = min(self.current_page * self.page_size, self.total_rows)
                self.page_info_label.setText(
                    f"第 {self.current_page}/{self.total_pages} 页 "
                    f"(显示 {start_row}-{end_row} 行，共 {self.total_rows} 行)"
                )
        else:
            # 客户端分页：使用 len(self.all_data)
            total_rows = len(self.all_data)
            if total_rows == 0:
                self.page_info_label.setText("第 0/0 页 (显示 0-0 行，共 0 行)")
            else:
                start_row = (self.current_page - 1) * self.page_size + 1
                end_row = min(self.current_page * self.page_size, total_rows)
                self.page_info_label.setText(
                    f"第 {self.current_page}/{self.total_pages} 页 "
                    f"(显示 {start_row}-{end_row} 行，共 {total_rows} 行)"
                )
    
    def _go_first_page(self):
        """跳转到首页"""
        if self.current_page != 1:
            self.current_page = 1
            self._display_current_page()
    
    def _go_prev_page(self):
        """上一页"""
        if self.current_page > 1:
            self.current_page -= 1
            self._display_current_page()
    
    def _go_next_page(self):
        """下一页"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self._display_current_page()
    
    def _go_last_page(self):
        """跳转到末页"""
        if self.current_page != self.total_pages:
            self.current_page = self.total_pages
            self._display_current_page()
    
    def _on_page_input(self):
        """处理页码输入"""
        try:
            page = int(self.page_input.text())
            if 1 <= page <= self.total_pages:
                self.current_page = page
                self._display_current_page()
            else:
                show_toast(self, f"页码必须在 1-{self.total_pages} 之间", "warning")
                self.page_input.setText(str(self.current_page))
        except ValueError:
            show_toast(self, "请输入有效的页码", "warning")
            self.page_input.setText(str(self.current_page))
    
    def _on_page_size_changed(self, new_size):
        """每页显示行数改变"""
        self.page_size = new_size
        
        # 重新计算总页数
        if self.server_side_paging:
            # 服务器端分页：使用 total_rows
            if self.total_rows > 0:
                self.total_pages = max(1, (self.total_rows + self.page_size - 1) // self.page_size)
                # 调整当前页码（如果超出范围）
                if self.current_page > self.total_pages:
                    self.current_page = self.total_pages
                # 重新加载当前页
                self._display_current_page()
        else:
            # 客户端分页：使用 len(self.all_data)
            if self.all_data:
                self.total_pages = max(1, (len(self.all_data) + self.page_size - 1) // self.page_size)
                # 调整当前页码（如果超出范围）
                if self.current_page > self.total_pages:
                    self.current_page = self.total_pages
                self._display_current_page()
    
    def _display_current_page(self):
        """显示当前页的数据"""
        if self.server_side_paging:
            # 服务器端分页：重新执行SQL
            self._load_page_from_server()
        else:
            # 客户端分页：从 all_data 中切分
            if not self.all_data:
                return
            
            # 计算当前页的数据范围
            start_idx = (self.current_page - 1) * self.page_size
            end_idx = min(start_idx + self.page_size, len(self.all_data))
            page_data = self.all_data[start_idx:end_idx]
            
            # 显示数据（不触发动画）
            self._fill_table_with_pagination(page_data)
            
            # 更新分页控件
            self._update_pagination_controls()
    
    def _start_count_query(self):
        """启动 COUNT 查询获取总行数"""
        if not self.connection_string or not self.original_sql:
            return
        
        from src.gui.workers.pagination_worker import PaginationWorker
        
        # 如果有正在运行的分页查询，先停止
        if self.pagination_worker and self.pagination_worker.isRunning():
            self.pagination_worker.stop()
            self.pagination_worker.wait(1000)
        
        # 创建分页 worker（只获取 COUNT，不查询数据）
        self.pagination_worker = PaginationWorker(
            self.connection_string,
            self.connect_args,
            self.original_sql,
            self.current_page,
            self.page_size,
            get_count=True  # 只获取 COUNT
        )
        
        # 连接信号
        self.pagination_worker.count_finished.connect(self._on_count_finished)
        
        # 启动线程
        self.pagination_worker.start()
    
    def _on_count_finished(self, total_rows: int):
        """COUNT 查询完成"""
        self.total_rows = total_rows
        self.total_pages = max(1, (total_rows + self.page_size - 1) // self.page_size)
        
        # 更新分页控件
        self._update_pagination_controls()
        
        logger.info(f"总行数: {total_rows}, 总页数: {self.total_pages}")
    
    def _load_page_from_server(self):
        """从服务器加载当前页数据"""
        if not self.connection_string or not self.original_sql:
            logger.warning("缺少连接信息或SQL，无法执行服务器端分页")
            return
        
        from src.gui.workers.pagination_worker import PaginationWorker
        
        # 如果有正在运行的分页查询，先停止
        if self.pagination_worker and self.pagination_worker.isRunning():
            self.pagination_worker.stop()
            self.pagination_worker.wait(1000)
        
        # 创建分页 worker
        self.pagination_worker = PaginationWorker(
            self.connection_string,
            self.connect_args,
            self.original_sql,
            self.current_page,
            self.page_size,
            get_count=False  # 不需要 COUNT（已经有了）
        )
        
        # 连接信号
        self.pagination_worker.query_finished.connect(self._on_page_loaded)
        
        # 启动线程
        self.pagination_worker.start()
        
        logger.info(f"加载第 {self.current_page} 页，每页 {self.page_size} 行")
    
    def _on_page_loaded(self, success: bool, data, error: str, columns):
        """分页数据加载完成"""
        if success and data is not None:
            # 更新表格数据
            self._fill_table_with_pagination(data)
            
            # 保存当前页数据
            self.all_data = data
            
            # 更新分页控件
            self._update_pagination_controls()
            
            logger.info(f"第 {self.current_page} 页加载完成: {len(data)} 行")
        else:
            # 显示错误
            logger.error(f"加载第 {self.current_page} 页失败: {error}")
            self._show_status_to_main_window(f"加载失败: {error}", 5000)
    
    def _show_status_to_main_window(self, message: str, timeout: int = 3000):
        """显示状态信息到主窗口状态栏"""
        if self.main_window and hasattr(self.main_window, 'statusBar'):
            self.main_window.statusBar().showMessage(message, timeout)
    
    def display_results(
        self, 
        data: List[Dict], 
        error: Optional[str] = None,
        affected_rows: Optional[int] = None,
        columns: Optional[List[str]] = None,
        connection_string: Optional[str] = None,
        connect_args: Optional[dict] = None
    ):
        """显示查询结果
        
        Args:
            data: 查询结果数据
            error: 错误信息
            affected_rows: 影响的行数
            columns: 列名列表
            connection_string: 数据库连接字符串（用于服务器端分页）
            connect_args: 连接参数（用于服务器端分页）
        """
        # 设置表格透明，给视觉反馈（仍然占位置）
        from PyQt6.QtWidgets import QGraphicsOpacityEffect
        opacity_effect = QGraphicsOpacityEffect()
        opacity_effect.setOpacity(0.3)  # 30% 不透明度
        self.table.setGraphicsEffect(opacity_effect)
        
        # 保存连接信息（用于服务器端分页）
        if connection_string:
            self.connection_string = connection_string
            self.connect_args = connect_args or {}
        if error:
            # 显示错误到主窗口状态栏
            self._show_status_to_main_window(f"错误: {error}", 5000)
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.raw_data = []
            self.all_data = []
            self.export_btn.setEnabled(False)
            self.pagination_widget.setVisible(False)
            # 恢复透明度
            self.table.setGraphicsEffect(None)
            return
        
        if affected_rows is not None:
            # 非查询语句
            self._show_status_to_main_window(f"执行成功: 影响 {affected_rows} 行")
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.raw_data = []
            self.all_data = []
            self.export_btn.setEnabled(False)
            self.pagination_widget.setVisible(False)
            # 恢复透明度
            self.table.setGraphicsEffect(None)
            return
        
        if not data:
            # 如果没有数据但有列名，显示表头
            if columns:
                self._show_status_to_main_window("查询完成: 0 行")
                self.table.setRowCount(0)
                self.table.setColumnCount(len(columns))
                self.table.setHorizontalHeaderLabels(columns)
                # 为每个表头添加提示（点击复制）
                for col_idx in range(len(columns)):
                    header_item = self.table.horizontalHeaderItem(col_idx)
                    if header_item:
                        header_item.setToolTip("点击复制列名")
                self.raw_data = []
                self.all_data = []
                self.export_btn.setEnabled(False)
                self.pagination_widget.setVisible(False)
                # 只在第一次显示或列数变化时调整列宽
                if not hasattr(self, '_last_column_count') or self._last_column_count != len(columns):
                    self._resize_columns_with_max_width()
                    self._last_column_count = len(columns)
            else:
                self._show_status_to_main_window("查询完成: 0 行")
                self.table.setRowCount(0)
                self.table.setColumnCount(0)
                self.raw_data = []
                self.all_data = []
                self.export_btn.setEnabled(False)
                self.pagination_widget.setVisible(False)
            # 恢复透明度
            self.table.setGraphicsEffect(None)
            return
        
        # 标记正在更新数据，避免触发itemChanged事件
        self._updating_data = True
        
        # 保存原始数据
        self.raw_data = data
        
        # 检测是否启用服务器端分页
        # 条件：1. 有连接信息  2. 有原始SQL  3. SQL是SELECT查询
        import re
        is_select_query = False
        if self.original_sql:
            sql_upper = self.original_sql.strip().upper()
            is_select_query = sql_upper.startswith("SELECT")
        
        if self.connection_string and self.original_sql and is_select_query:
            # 启用服务器端分页
            self.server_side_paging = True
            
            # 移除SQL中的LIMIT子句（如果有）
            sql_no_limit = re.sub(r'\s+LIMIT\s+\d+(\s+OFFSET\s+\d+)?', '', self.original_sql, flags=re.IGNORECASE)
            self.original_sql = sql_no_limit.strip().rstrip(';')
            
            # 当前页数据
            self.all_data = data
            self.current_page = 1
            
            # 启动 COUNT 查询获取总行数
            self._start_count_query()
            
            # 暂时使用当前数据量计算总页数（COUNT完成后会更新）
            self.total_rows = len(data)
            self.total_pages = max(1, (len(data) + self.page_size - 1) // self.page_size)
            
            logger.info(f"启用服务器端分页: page_size={self.page_size}")
        else:
            # 使用客户端分页（数据已全部加载）
            self.server_side_paging = False
            self.all_data = data
            self.current_page = 1
            self.total_rows = len(data)
            
            # 计算总页数
            self.total_pages = max(1, (len(data) + self.page_size - 1) // self.page_size)
        
        # 显示分页控件（数据超过10行时显示）
        self.pagination_widget.setVisible(len(data) > 10)
        
        # 保存原始数据的副本（用于生成WHERE条件）
        import copy
        self.original_data = copy.deepcopy(data)
        
        # 清空修改记录
        self.modified_cells.clear()
        
        # 启用导出按钮（只有在有数据时才启用）
        if data and len(data) > 0:
            self.export_btn.setEnabled(True)
        else:
            self.export_btn.setEnabled(False)
        
        # 显示第一页数据
        page_data = data[:min(self.page_size, len(data))]
        
        # 显示数据
        columns = list(data[0].keys())
        self.table.setRowCount(len(page_data))
        self.table.setColumnCount(len(columns))
        
        self.table.setHorizontalHeaderLabels(columns)
        
        # 为每个表头添加提示（点击复制）
        for col_idx in range(len(columns)):
            header_item = self.table.horizontalHeaderItem(col_idx)
            if header_item:
                header_item.setToolTip("点击复制列名")
        
        # 填充数据
        for row_idx, row_data in enumerate(page_data):
            for col_idx, col_name in enumerate(columns):
                value = row_data.get(col_name)
                
                # 处理None值
                if value is None:
                    display_value = "NULL"
                else:
                    display_value = str(value)
                
                item = QTableWidgetItem(display_value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                
                # NULL值特殊样式
                if value is None:
                    item.setForeground(Qt.GlobalColor.gray)
                
                # 设置单元格可编辑（包括NULL值）
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                
                self.table.setItem(row_idx, col_idx, item)
        
        # 只在第一次显示或列数变化时调整列宽
        if not hasattr(self, '_last_column_count') or self._last_column_count != len(columns):
            self._resize_columns_with_max_width()
            self._last_column_count = len(columns)
        
        # 更新分页控件
        if self.pagination_widget.isVisible():
            self._update_pagination_controls()
        
        # 更新状态到主窗口状态栏
        total_rows = len(data)
        if total_rows <= self.page_size:
            self._show_status_to_main_window(f"查询完成: {total_rows} 行, {len(columns)} 列")
        else:
            self._show_status_to_main_window(f"查询完成: 共 {total_rows} 行，显示前 {min(self.page_size, total_rows)} 行")
        
        # 数据更新完成
        self._updating_data = False
        
        # 延迟0.2秒后恢复表格透明度
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(200, self._restore_table_opacity)
    
    def _fill_table_with_pagination(self, page_data: List[Dict]):
        """填充表格数据（用于分页切换）"""
        if not page_data:
            return
        
        # 标记正在更新数据，避免触发itemChanged事件
        self._updating_data = True
        
        columns = list(page_data[0].keys())
        self.table.setRowCount(len(page_data))
        
        # 填充数据
        for row_idx, row_data in enumerate(page_data):
            for col_idx, col_name in enumerate(columns):
                value = row_data.get(col_name)
                
                # 处理None值
                if value is None:
                    display_value = "NULL"
                else:
                    display_value = str(value)
                
                item = QTableWidgetItem(display_value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                
                # NULL值特殊样式
                if value is None:
                    item.setForeground(Qt.GlobalColor.gray)
                
                # 设置单元格可编辑（包括NULL值）
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                
                self.table.setItem(row_idx, col_idx, item)
        
        # 分页切换时不调整列宽，保持用户设置的列宽
        # （列宽只在首次显示时调整）
        
        # 数据更新完成
        self._updating_data = False
        
        # 延迟0.2秒后恢复表格透明度
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(200, self._restore_table_opacity)
    
    def _resize_columns_with_max_width(self):
        """调整列宽，限制最小和最大宽度"""
        # 先根据内容调整列宽
        self.table.resizeColumnsToContents()
        
        # 设置最小和最大宽度限制
        min_column_width = 80   # 最小宽度：80像素
        max_column_width = 400  # 最大宽度：400像素
        
        # 限制每列的宽度在最小和最大值之间
        header = self.table.horizontalHeader()
        for col_idx in range(self.table.columnCount()):
            current_width = header.sectionSize(col_idx)
            
            if current_width < min_column_width:
                header.resizeSection(col_idx, min_column_width)
            elif current_width > max_column_width:
                header.resizeSection(col_idx, max_column_width)
    
    def on_header_clicked(self, logical_index: int):
        """表头点击事件：复制列名到剪贴板"""
        header_item = self.table.horizontalHeaderItem(logical_index)
        if header_item:
            column_name = header_item.text()
            clipboard = QApplication.clipboard()
            clipboard.setText(column_name)
            # 显示Toast提示
            show_toast(f"已复制列名: {column_name}", parent=self.table, duration=2000)
    
    def on_row_double_clicked(self, item: QTableWidgetItem):
        """双击单元格时进入编辑模式"""
        # 双击时直接进入编辑模式（默认行为）
        pass
    
    def show_context_menu(self, position):
        """显示右键菜单"""
        item = self.table.itemAt(position)
        
        # 获取选中的单元格和行
        selected_items = self.table.selectedItems()
        selected_rows = sorted({it.row() for it in selected_items}) if selected_items else []
        
        # 创建右键菜单
        menu = QMenu(self)
        
        # 添加"刷新"选项（如果有原始SQL）
        if self.original_sql and self.main_window:
            refresh_action = menu.addAction("🔄 刷新")
            refresh_action.triggered.connect(self.refresh_data)
            menu.addSeparator()
        
        # 如果有选中的单元格，添加其他选项
        if item:
            row = item.row()
            if 0 <= row < len(self.raw_data):
                # 添加"查看 JSON 数据"选项
                json_action = menu.addAction("查看 JSON 数据")
                json_action.triggered.connect(lambda: self.show_json_dialog(row))

                # 如果有选中的单元格，添加"填充为 NULL"选项
                if selected_items:
                    fill_null_action = menu.addAction("设置为NULL")
                    fill_null_action.triggered.connect(self.fill_selected_cells_with_null)
                
                # 如果有选中的行，添加删除选项
                if selected_rows:
                    menu.addSeparator()
                    delete_action = menu.addAction(f"删除选中行 ({len(selected_rows)} 行)")
                    delete_action.triggered.connect(lambda: self.delete_selected_rows(selected_rows))
        
        # 显示菜单
        menu.exec(self.table.mapToGlobal(position))
    
    def refresh_data(self):
        """刷新数据：重新执行原始SQL查询"""
        if not self.original_sql:
            return
        
        # 显示刷新状态
        self._show_status_to_main_window("正在刷新数据...", timeout=0)
        
        # 优先使用自定义的执行查询函数（新标签）
        if self.execute_query_func:
            self.execute_query_func(self.original_sql)
        # 否则使用主窗口的执行查询方法（第一个查询标签）
        elif self.main_window and hasattr(self.main_window, 'execute_query'):
            self.main_window.execute_query(self.original_sql)
        else:
            self._show_status_to_main_window("无法刷新：缺少执行查询函数", timeout=3000)

    def fill_selected_cells_with_null(self):
        """将选中的单元格填充为 NULL（文本为 'NULL'，触发现有更新逻辑）"""
        items = self.table.selectedItems()
        if not items:
            return

        for it in items:
            # 设置显示文本为 "NULL"；on_item_changed 会把它转换为 None 并更新数据库
            it.setText("NULL")
    
    def show_json_dialog(self, row: int):
        """显示JSON数据对话框"""
        if not self.raw_data:
            return
        
        if row < 0 or row >= len(self.raw_data):
            return
        
        # 获取该行的数据
        row_data = self.raw_data[row]
        
        # 创建JSON显示对话框
        dialog = QDialog(self)
        dialog.setWindowTitle(f"行 {row + 1} 的JSON数据")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout()
        dialog.setLayout(layout)
        
        # JSON文本编辑器
        json_edit = QTextEdit()
        json_edit.setReadOnly(True)
        json_edit.setFont(QFont("Consolas", 10))
        
        # 格式化JSON（处理datetime等特殊类型）
        try:
            # 自定义JSON编码器，处理datetime等类型
            class CustomJSONEncoder(json.JSONEncoder):
                def default(self, obj):
                    if isinstance(obj, (datetime, date)):
                        return obj.isoformat()
                    elif isinstance(obj, time):
                        return obj.isoformat()
                    elif isinstance(obj, Decimal):
                        return float(obj)
                    elif hasattr(obj, '__dict__'):
                        return obj.__dict__
                    return super().default(obj)
            
            json_text = json.dumps(row_data, ensure_ascii=False, indent=2, cls=CustomJSONEncoder)
        except Exception as e:
            json_text = f"无法序列化为JSON: {str(e)}\n\n原始数据:\n{str(row_data)}"
        
        json_edit.setPlainText(json_text)
        
        # 设置JSON语法高亮
        highlighter = JSONHighlighter(json_edit.document())
        
        layout.addWidget(json_edit)
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        button_box.accepted.connect(dialog.accept)
        layout.addWidget(button_box)
        
        dialog.exec()
    
    def on_item_changed(self, item: QTableWidgetItem):
        """单元格内容改变时的回调"""
        # 如果正在更新数据，忽略此事件
        if self._updating_data:
            return
        
        # 获取行和列
        row = item.row()
        col = item.column()
        
        # 更新原始数据
        if self.raw_data and 0 <= row < len(self.raw_data):
            columns = list(self.raw_data[0].keys())
            if 0 <= col < len(columns):
                col_name = columns[col]
                new_value = item.text()
                
                # 获取原始值
                old_value = self.raw_data[row].get(col_name)
                old_display_value = "NULL" if old_value is None else str(old_value)
                
                # 处理NULL值
                if new_value.upper() == "NULL" or new_value == "":
                    self.raw_data[row][col_name] = None
                    item.setForeground(Qt.GlobalColor.gray)
                    new_value_for_db = None
                else:
                    self.raw_data[row][col_name] = new_value
                    item.setForeground(Qt.GlobalColor.black)
                    new_value_for_db = new_value
                
                # 记录修改（如果值确实改变了）
                if old_display_value != new_value:
                    self.modified_cells[(row, col)] = (old_value, new_value_for_db)
                    # 标记单元格为已修改（可选：改变背景色）
                    item.setBackground(QColor(255, 255, 200))  # 浅黄色背景表示已修改
                    
                    # 自动更新到数据库
                    self._update_to_database(row, col_name, new_value_for_db, old_value)
                else:
                    # 如果值没有改变，移除修改记录
                    if (row, col) in self.modified_cells:
                        del self.modified_cells[(row, col)]
                        # 恢复默认背景（使用透明或白色）
                        item.setBackground(QColor(255, 255, 255, 0))  # 透明背景，让系统样式生效
                
                # 更新状态到主窗口状态栏（可选：显示已修改标记）
                if self.modified_cells:
                    self._show_status_to_main_window(f"查询完成: {len(self.raw_data)} 行, {len(columns)} 列 (已修改 {len(self.modified_cells)} 个单元格)")
                else:
                    self._show_status_to_main_window(f"查询完成: {len(self.raw_data)} 行, {len(columns)} 列")
    
    def _update_to_database(self, row: int, col_name: str, new_value, old_value):
        """更新单元格到数据库"""
        if not self.main_window or not self.original_sql:
            return
        
        # 检查是否是SELECT查询
        sql_upper = self.original_sql.strip().upper()
        if not sql_upper.startswith("SELECT"):
            return
        
        # 从原始SQL中提取表名
        table_name = self._extract_table_name_from_sql(self.original_sql)
        if not table_name:
            # 如果无法提取表名，尝试从SQL中提取
            return
        
        # 获取该行的原始数据（用于WHERE条件）
        if row < 0 or row >= len(self.original_data):
            return
        
        original_row_data = self.original_data[row]
        columns = list(original_row_data.keys())
        
        # 生成UPDATE SQL语句
        # 使用所有列的值作为WHERE条件（这样可以唯一标识一行）
        update_sql = self._generate_update_sql(table_name, col_name, new_value, original_row_data, columns)
        
        if update_sql:
            # 执行UPDATE语句
            self._execute_update(update_sql)
    
    def _extract_table_name_from_sql(self, sql: str) -> Optional[str]:
        """从SQL中提取表名"""
        import re
        sql_upper = sql.strip().upper()
        
        # 只处理SELECT查询
        if not sql_upper.startswith("SELECT"):
            return None
        
        # 更精确的正则表达式，处理反引号和点号
        # 需要处理的情况：
        # 1. FROM `database`.`table` - 标准格式
        # 2. FROM `database.table`.`table` - 数据库名包含点号
        # 3. FROM database.table - 不带反引号
        # 4. FROM `table` - 只有表名
        
        # 先尝试匹配带反引号的格式：FROM `xxx`.`yyy`
        # 这个模式会匹配最后一个点号分隔的两部分
        pattern1 = r'FROM\s+`([^`]+)`\.`([^`]+)`'
        match = re.search(pattern1, sql, re.IGNORECASE)
        if match:
            # 第一部分可能是 database 或 database.table（数据库名包含点号）
            # 第二部分是表名
            db_part = match.group(1).strip()
            table_name = match.group(2).strip()
            # 返回 database.table 格式（完整路径）
            return f"{db_part}.{table_name}"
        
        # 尝试匹配不带反引号的格式：FROM xxx.yyy
        pattern2 = r'FROM\s+([^\s`]+)\.([^\s`]+)'
        match = re.search(pattern2, sql, re.IGNORECASE)
        if match:
            db_part = match.group(1).strip()
            table_name = match.group(2).strip()
            # 移除可能的反引号
            db_part = db_part.strip('`')
            table_name = table_name.strip('`')
            return f"{db_part}.{table_name}"
        
        # 尝试匹配单个表名（带反引号）
        pattern3 = r'FROM\s+`([^`]+)`'
        match = re.search(pattern3, sql, re.IGNORECASE)
        if match:
            table_name = match.group(1).strip()
            return table_name
        
        # 尝试匹配单个表名（不带反引号）
        pattern4 = r'FROM\s+([^\s`]+)'
        match = re.search(pattern4, sql, re.IGNORECASE)
        if match:
            table_name = match.group(1).strip()
            # 移除可能的反引号
            table_name = table_name.strip('`')
            return table_name
        
        return None
    
    def _get_primary_keys(self, table_name: str) -> List[str]:
        """获取表的主键列名列表（优先使用主键过滤）"""
        if not self.main_window or not hasattr(self.main_window, 'db_manager'):
            return []
        
        try:
            connection_id = getattr(self.main_window, "current_connection_id", None)
            if not connection_id:
                return []
            
            db_manager = self.main_window.db_manager
            engine = db_manager.get_engine(connection_id)
            if not engine:
                return []
            
            connection = db_manager.get_connection(connection_id)
            current_db = connection.database if connection else None
            db_type = connection.db_type.value if connection and connection.db_type else ""
            
            from sqlalchemy import inspect
            inspector = inspect(engine)
            
            # 解析表名和 schema
            actual_table = table_name.strip().strip('`')
            schema_name = None
            if '.' in actual_table:
                last_dot = actual_table.rfind('.')
                schema_name = actual_table[:last_dot].strip().strip('`')
                actual_table = actual_table[last_dot + 1:].strip().strip('`')
            elif current_db:
                schema_name = current_db
            
            # 使用 get_pk_constraint 兼容不同数据库/版本
            if schema_name and db_type in ("mysql", "mariadb", "postgresql"):
                pk_constraint = inspector.get_pk_constraint(actual_table, schema=schema_name)
            else:
                pk_constraint = inspector.get_pk_constraint(actual_table)
            
            primary_keys = pk_constraint.get("constrained_columns", []) if pk_constraint else []
            if not primary_keys:
                logger.warning(f"未获取到主键: table={table_name}, schema={schema_name}")
            else:
                logger.info(f"主键列: {primary_keys} (table={table_name}, schema={schema_name})")
            return primary_keys
        except Exception as e:
            logger.debug(f"获取表 {table_name} 的主键失败: {str(e)}")
            return []
    
    def _generate_update_sql(self, table_name: str, col_name: str, new_value, original_row_data: Dict, columns: List[str]) -> Optional[str]:
        """生成UPDATE SQL语句"""
        # 获取数据库类型
        from src.core.database_connection import DatabaseType
        db_type = None
        if self.main_window and hasattr(self.main_window, 'current_connection_id'):
            connection = self.main_window.db_manager.get_connection(self.main_window.current_connection_id)
            if connection:
                db_type = connection.db_type
        
        # 转义表名和列名（根据数据库类型使用不同的引用符号）
        def escape_identifier(name: str) -> str:
            # 先移除所有可能的引用符号
            name = name.strip().strip('`').strip('"').strip('[').strip(']')
            
            # 根据数据库类型选择引用符号
            if db_type in (DatabaseType.MYSQL, DatabaseType.MARIADB):
                quote_char = '`'
            elif db_type in (DatabaseType.POSTGRESQL, DatabaseType.SQLITE):
                quote_char = '"'
            elif db_type == DatabaseType.SQLSERVER:
                # SQL Server 使用方括号
                quote_start = '['
                quote_end = ']'
            else:
                # 其他数据库类型不使用引用符号
                return name
            
            # 处理 SQL Server 的特殊情况
            if db_type == DatabaseType.SQLSERVER:
                if '.' in name:
                    last_dot_index = name.rfind('.')
                    db_part = name[:last_dot_index].strip()
                    table_part = name[last_dot_index + 1:].strip()
                    if db_part and table_part:
                        return f"{quote_start}{db_part}{quote_end}.{quote_start}{table_part}{quote_end}"
                    elif table_part:
                        return f"{quote_start}{table_part}{quote_end}"
                return f"{quote_start}{name}{quote_end}" if name else name
            
            # 处理带点号的标识符（database.table）
            if '.' in name:
                last_dot_index = name.rfind('.')
                db_part = name[:last_dot_index].strip()
                table_part = name[last_dot_index + 1:].strip()
                if db_part and table_part:
                    return f"{quote_char}{db_part}{quote_char}.{quote_char}{table_part}{quote_char}"
                elif table_part:
                    return f"{quote_char}{table_part}{quote_char}"
            
            # 单个标识符
            return f"{quote_char}{name}{quote_char}" if name else name
        
        # 转义值（处理SQL注入和JSON字段）
        def escape_value(value) -> str:
            if value is None:
                return "NULL"
            elif isinstance(value, str):
                # 转义字符串中的特殊字符（用于JSON字段等）
                # 1. 先转义反斜杠（必须在其他转义之前）
                escaped = value.replace("\\", "\\\\")
                # 2. 转义单引号
                escaped = escaped.replace("'", "''")
                # 3. 转义换行符、回车符、制表符等控制字符
                escaped = escaped.replace("\n", "\\n")
                escaped = escaped.replace("\r", "\\r")
                escaped = escaped.replace("\t", "\\t")
                escaped = escaped.replace("\0", "\\0")
                return f"'{escaped}'"
            elif isinstance(value, (int, float)):
                return str(value)
            elif isinstance(value, bool):
                return "1" if value else "0"
            else:
                # 其他类型转为字符串
                str_value = str(value)
                # 转义特殊字符
                escaped = str_value.replace("\\", "\\\\")
                escaped = escaped.replace("'", "''")
                escaped = escaped.replace("\n", "\\n")
                escaped = escaped.replace("\r", "\\r")
                escaped = escaped.replace("\t", "\\t")
                escaped = escaped.replace("\0", "\\0")
                return f"'{escaped}'"
        
        # 构建SET子句
        set_clause = f"{escape_identifier(col_name)} = {escape_value(new_value)}"
        
        # 获取主键列
        primary_keys = self._get_primary_keys(table_name)
        
        # 构建WHERE子句
        # 如果有主键，优先使用主键；否则使用所有列
        where_columns = primary_keys if primary_keys else columns
        
        where_conditions = []
        for col in where_columns:
            # 确保列在原始数据中存在
            if col not in original_row_data:
                continue
            value = original_row_data.get(col)
            if value is None:
                where_conditions.append(f"{escape_identifier(col)} IS NULL")
            else:
                where_conditions.append(f"{escape_identifier(col)} = {escape_value(value)}")
        
        # 如果没有有效的WHERE条件，回退到使用所有列
        if not where_conditions:
            where_conditions = []
            for col in columns:
                value = original_row_data.get(col)
                if value is None:
                    where_conditions.append(f"{escape_identifier(col)} IS NULL")
                else:
                    where_conditions.append(f"{escape_identifier(col)} = {escape_value(value)}")
        
        where_clause = " AND ".join(where_conditions)
        
        # 生成完整的UPDATE语句
        update_sql = f"UPDATE {escape_identifier(table_name)} SET {set_clause} WHERE {where_clause}"
        
        return update_sql
    
    def _execute_update(self, update_sql: str):
        """执行UPDATE语句"""
        if not self.main_window:
            return
        
        try:
            # 在主窗口状态栏显示SQL
            self._show_status_to_main_window(f"执行UPDATE: {update_sql}", 5000)
            
            # 停止之前的UPDATE worker（如果存在）
            if self.update_worker and self.update_worker.isRunning():
                self.update_worker.stop()
                self.update_worker.wait(1000)
                if self.update_worker.isRunning():
                    self.update_worker.terminate()
                    self.update_worker.wait(500)
                try:
                    self.update_worker.query_finished.disconnect()
                except:
                    pass
                self.update_worker.deleteLater()
            
            # 使用主窗口的execute_query方法执行UPDATE
            # 注意：这里直接执行，不显示在SQL编辑器中
            from src.gui.workers.query_worker import QueryWorker
            
            connection = self.main_window.db_manager.get_connection(self.main_window.current_connection_id)
            if not connection:
                return
            
            # 创建并启动工作线程执行UPDATE
            self.update_worker = QueryWorker(
                connection.get_connection_string(),
                connection.get_connect_args(),
                update_sql,
                is_query=False  # UPDATE不是查询
            )
            
            # 连接信号
            self.update_worker.query_finished.connect(
                lambda success, data, error, affected_rows: self._on_update_finished(success, error, affected_rows)
            )
            
            # 启动线程
            self.update_worker.start()
            
        except Exception as e:
            logger.error(f"执行UPDATE失败: {str(e)}")
            QMessageBox.warning(self, "更新失败", f"更新数据库失败: {str(e)}")
    
    def _on_update_finished(self, success: bool, error: Optional[str], affected_rows: Optional[int]):
        """UPDATE执行完成回调"""
        if success:
            # 更新成功，更新原始数据
            import copy
            self.original_data = copy.deepcopy(self.raw_data)
            # 清空修改记录
            self.modified_cells.clear()
            # 恢复所有单元格的背景色
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item:
                        # 恢复默认背景（使用透明，让系统样式生效）
                        item.setBackground(QColor(255, 255, 255, 0))  # 透明背景
            
            self._show_status_to_main_window(f"查询完成: {len(self.raw_data)} 行 (已保存到数据库)")
        else:
            QMessageBox.warning(self, "更新失败", f"更新数据库失败: {error}")
    
    def delete_selected_rows(self, selected_rows: List[int]):
        """删除选中的行"""
        if not selected_rows:
            return
        
        if not self.main_window or not self.original_sql:
            QMessageBox.warning(self, "警告", "无法删除：缺少SQL信息")
            return
        
        # 检查是否是SELECT查询
        sql_upper = self.original_sql.strip().upper()
        if not sql_upper.startswith("SELECT"):
            QMessageBox.warning(self, "警告", "只能删除SELECT查询的结果")
            return
        
        # 从原始SQL中提取表名
        table_name = self._extract_table_name_from_sql(self.original_sql)
        if not table_name:
            QMessageBox.warning(self, "警告", "无法从SQL中提取表名")
            return
        
        # 确认删除
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除选中的 {len(selected_rows)} 行数据吗？\n\n此操作不可撤销！",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # 生成DELETE SQL语句
        delete_sqls = []
        columns = list(self.raw_data[0].keys()) if self.raw_data else []
        
        for row_idx in selected_rows:
            if row_idx < 0 or row_idx >= len(self.original_data):
                continue
            
            original_row_data = self.original_data[row_idx]
            delete_sql = self._generate_delete_sql(table_name, original_row_data, columns)
            if delete_sql:
                delete_sqls.append(delete_sql)
        
        if not delete_sqls:
            QMessageBox.warning(self, "警告", "无法生成DELETE语句")
            return
        
        # 执行DELETE语句
        self._execute_delete(delete_sqls, selected_rows)
    
    def _generate_delete_sql(self, table_name: str, original_row_data: Dict, columns: List[str]) -> Optional[str]:
        """生成DELETE SQL语句"""
        # 获取数据库类型
        from src.core.database_connection import DatabaseType
        db_type = None
        if self.main_window and hasattr(self.main_window, 'current_connection_id'):
            connection = self.main_window.db_manager.get_connection(self.main_window.current_connection_id)
            if connection:
                db_type = connection.db_type
        
        # 转义表名（根据数据库类型使用不同的引用符号）
        def escape_identifier(name: str) -> str:
            # 先移除所有可能的引用符号
            name = name.strip().strip('`').strip('"').strip('[').strip(']')
            
            # 根据数据库类型选择引用符号
            if db_type in (DatabaseType.MYSQL, DatabaseType.MARIADB):
                quote_char = '`'
            elif db_type in (DatabaseType.POSTGRESQL, DatabaseType.SQLITE):
                quote_char = '"'
            elif db_type == DatabaseType.SQLSERVER:
                # SQL Server 使用方括号
                quote_start = '['
                quote_end = ']'
            else:
                # 其他数据库类型不使用引用符号
                return name
            
            # 处理 SQL Server 的特殊情况
            if db_type == DatabaseType.SQLSERVER:
                if '.' in name:
                    last_dot_index = name.rfind('.')
                    db_part = name[:last_dot_index].strip()
                    table_part = name[last_dot_index + 1:].strip()
                    if db_part and table_part:
                        return f"{quote_start}{db_part}{quote_end}.{quote_start}{table_part}{quote_end}"
                    elif table_part:
                        return f"{quote_start}{table_part}{quote_end}"
                return f"{quote_start}{name}{quote_end}" if name else name
            
            # 处理带点号的标识符（database.table）
            if '.' in name:
                last_dot_index = name.rfind('.')
                db_part = name[:last_dot_index].strip()
                table_part = name[last_dot_index + 1:].strip()
                if db_part and table_part:
                    return f"{quote_char}{db_part}{quote_char}.{quote_char}{table_part}{quote_char}"
                elif table_part:
                    return f"{quote_char}{table_part}{quote_char}"
            
            # 单个标识符
            return f"{quote_char}{name}{quote_char}" if name else name
        
        # 转义值（处理SQL注入和JSON字段）
        def escape_value(value) -> str:
            if value is None:
                return "NULL"
            elif isinstance(value, str):
                # 转义字符串中的特殊字符（用于JSON字段等）
                # 1. 先转义反斜杠（必须在其他转义之前）
                escaped = value.replace("\\", "\\\\")
                # 2. 转义单引号
                escaped = escaped.replace("'", "''")
                # 3. 转义换行符、回车符、制表符等控制字符
                escaped = escaped.replace("\n", "\\n")
                escaped = escaped.replace("\r", "\\r")
                escaped = escaped.replace("\t", "\\t")
                escaped = escaped.replace("\0", "\\0")
                return f"'{escaped}'"
            elif isinstance(value, (int, float)):
                return str(value)
            elif isinstance(value, bool):
                return "1" if value else "0"
            else:
                # 其他类型转为字符串
                str_value = str(value)
                # 转义特殊字符
                escaped = str_value.replace("\\", "\\\\")
                escaped = escaped.replace("'", "''")
                escaped = escaped.replace("\n", "\\n")
                escaped = escaped.replace("\r", "\\r")
                escaped = escaped.replace("\t", "\\t")
                escaped = escaped.replace("\0", "\\0")
                return f"'{escaped}'"
        
        # 获取主键列
        primary_keys = self._get_primary_keys(table_name)
        
        # 构建WHERE子句
        # 如果有主键，优先使用主键；否则使用所有列
        where_columns = primary_keys if primary_keys else columns
        
        where_conditions = []
        for col in where_columns:
            # 确保列在原始数据中存在
            if col not in original_row_data:
                continue
            value = original_row_data.get(col)
            if value is None:
                where_conditions.append(f"{escape_identifier(col)} IS NULL")
            else:
                where_conditions.append(f"{escape_identifier(col)} = {escape_value(value)}")
        
        # 如果没有有效的WHERE条件，回退到使用所有列
        if not where_conditions:
            where_conditions = []
            for col in columns:
                value = original_row_data.get(col)
                if value is None:
                    where_conditions.append(f"{escape_identifier(col)} IS NULL")
                else:
                    where_conditions.append(f"{escape_identifier(col)} = {escape_value(value)}")
        
        where_clause = " AND ".join(where_conditions)
        
        # 生成完整的DELETE语句
        delete_sql = f"DELETE FROM {escape_identifier(table_name)} WHERE {where_clause}"
        
        return delete_sql
    
    def _execute_delete(self, delete_sqls: List[str], selected_rows: List[int]):
        """执行DELETE语句"""
        if not self.main_window:
            return
        
        try:
            # 合并多个DELETE语句（如果数据库支持）
            combined_sql = ";\n".join(delete_sqls)
            
            # 在日志中打印DELETE SQL
            import logging
            logger = logging.getLogger(__name__)
            logger.info("=" * 80)
            logger.info("执行DELETE语句:")
            for idx, sql in enumerate(delete_sqls, 1):
                logger.info(f"DELETE {idx}: {sql}")
            logger.info("=" * 80)
            
            # 在主窗口状态栏显示SQL
            # 如果SQL太长，只显示前200个字符
            display_sql = combined_sql[:200] + "..." if len(combined_sql) > 200 else combined_sql
            self._show_status_to_main_window(f"执行DELETE: {display_sql}", 5000)
            
            # 停止之前的UPDATE worker（如果存在）
            if self.update_worker and self.update_worker.isRunning():
                self.update_worker.stop()
                self.update_worker.wait(1000)
                if self.update_worker.isRunning():
                    self.update_worker.terminate()
                    self.update_worker.wait(500)
                try:
                    self.update_worker.query_finished.disconnect()
                except:
                    pass
                self.update_worker.deleteLater()
            
            from src.gui.workers.query_worker import QueryWorker
            
            connection = self.main_window.db_manager.get_connection(self.main_window.current_connection_id)
            if not connection:
                QMessageBox.warning(self, "警告", "数据库连接不存在")
                return
            
            # 创建并启动工作线程执行DELETE
            self.update_worker = QueryWorker(
                connection.get_connection_string(),
                connection.get_connect_args(),
                combined_sql,
                is_query=False  # DELETE不是查询
            )
            
            # 连接信号（支持单条和多条SQL）
            self.update_worker.query_finished.connect(
                lambda success, data, error, affected_rows, columns: self._on_delete_finished(
                    success, error, affected_rows, selected_rows
                )
            )
            # 如果有多条DELETE语句，使用multi_query_finished信号
            self.update_worker.multi_query_finished.connect(
                lambda results: self._on_multi_delete_finished(results, selected_rows)
            )
            
            # 启动线程
            self.update_worker.start()
            
        except Exception as e:
            logger.error(f"执行DELETE失败: {str(e)}")
            QMessageBox.warning(self, "删除失败", f"删除数据失败: {str(e)}")
    
    def _on_delete_finished(self, success: bool, error: Optional[str], affected_rows: Optional[int], selected_rows: List[int]):
        """DELETE执行完成回调（单条SQL）"""
        if success:
            # 删除成功，从表格中移除行（从后往前删除，避免索引变化）
            self._remove_rows_from_table(selected_rows)
            
            # 更新状态到主窗口状态栏
            remaining_rows = len(self.raw_data)
            self._show_status_to_main_window(f"删除成功: 已删除 {len(selected_rows)} 行，剩余 {remaining_rows} 行")
            
            QMessageBox.information(self, "删除成功", f"已成功删除 {len(selected_rows)} 行数据")
        else:
            self._show_status_to_main_window(f"查询完成: {len(self.raw_data)} 行")
            QMessageBox.warning(self, "删除失败", f"删除数据失败: {error}")
    
    def _on_multi_delete_finished(self, results: List[tuple], selected_rows: List[int]):
        """多条DELETE执行完成回调"""
        # results格式: [(sql, success, data, error, affected_rows, columns), ...]
        success_count = sum(1 for r in results if r[1])  # 统计成功的数量
        error_count = len(results) - success_count
        
        if success_count > 0:
            # 至少有一条删除成功，从表格中移除行
            self._remove_rows_from_table(selected_rows)
            
            # 更新状态到主窗口状态栏
            remaining_rows = len(self.raw_data)
            if error_count > 0:
                self._show_status_to_main_window(f"删除完成: 成功 {success_count} 行，失败 {error_count} 行，剩余 {remaining_rows} 行")
                QMessageBox.warning(self, "删除部分成功", f"成功删除 {success_count} 行，失败 {error_count} 行")
            else:
                self._show_status_to_main_window(f"删除成功: 已删除 {success_count} 行，剩余 {remaining_rows} 行")
                QMessageBox.information(self, "删除成功", f"已成功删除 {success_count} 行数据")
        else:
            # 全部失败
            error_messages = [r[3] for r in results if r[3]]
            error_msg = error_messages[0] if error_messages else "未知错误"
            self._show_status_to_main_window(f"查询完成: {len(self.raw_data)} 行")
            QMessageBox.warning(self, "删除失败", f"删除数据失败: {error_msg}")
    
    def _remove_rows_from_table(self, selected_rows: List[int]):
        """从表格中移除指定的行（从后往前删除，避免索引变化）"""
        for row_idx in reversed(sorted(selected_rows)):
            if 0 <= row_idx < len(self.raw_data):
                # 从数据中移除
                self.raw_data.pop(row_idx)
                self.original_data.pop(row_idx)
                # 从表格中移除
                self.table.removeRow(row_idx)
                # 更新修改记录中的行号（如果有）
                keys_to_update = []
                for (r, c), (old_val, new_val) in list(self.modified_cells.items()):
                    if r > row_idx:
                        keys_to_update.append((r, c, old_val, new_val))
                for r, c, old_val, new_val in keys_to_update:
                    del self.modified_cells[(r, c)]
                    self.modified_cells[(r - 1, c)] = (old_val, new_val)
    
    def show_export_menu(self):
        """显示导出菜单"""
        menu = QMenu(self)
        
        csv_action = menu.addAction("导出为 CSV")
        csv_action.triggered.connect(lambda: self.export_to_csv())
        
        excel_action = menu.addAction("导出为 Excel")
        excel_action.triggered.connect(lambda: self.export_to_excel())
        
        # 显示菜单
        button_pos = self.export_btn.mapToGlobal(self.export_btn.rect().bottomLeft())
        menu.exec(button_pos)
    
    def export_to_csv(self):
        """导出为CSV"""
        if not self.raw_data:
            QMessageBox.warning(self, "警告", "没有数据可导出")
            return
        
        # 检查是否只显示了部分数据
        should_fetch_all = False
        
        # 如果自动添加了LIMIT，提示用户是否导出全部数据
        if self.auto_limit_added or self.server_side_paging:
            reply = QMessageBox.question(
                self,
                "导出选项",
                f"当前显示 {len(self.raw_data)} 行数据（已自动限制显示）。\n\n"
                f"是否要重新查询并导出全部数据？\n\n"
                f"• 是：执行完整查询，导出所有数据（后台流式处理）\n"
                f"• 否：仅导出当前显示的 {len(self.raw_data)} 行\n\n"
                f"⚠️ 提示：导出全部数据会在后台进行，不会卡住界面",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Yes
            )
            if reply == QMessageBox.StandardButton.Cancel:
                return
            should_fetch_all = (reply == QMessageBox.StandardButton.Yes)
        
        # 选择保存文件
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出为CSV",
            "",
            "CSV文件 (*.csv);;所有文件 (*)"
        )
        
        if not file_path:
            return
        
        # 如果只导出当前数据，使用快速方法
        if not should_fetch_all:
            self._export_current_data_to_csv(file_path)
            return
        
        # 使用后台Worker导出全部数据
        self._start_background_export(file_path, 'csv')
    
    def _export_current_data_to_csv(self, file_path: str):
        """导出当前显示的数据到CSV（同步方法）"""
        try:
            columns = list(self.raw_data[0].keys())
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                
                for row in self.raw_data:
                    processed_row = {}
                    for key, value in row.items():
                        if isinstance(value, (datetime, date, time)):
                            processed_row[key] = value.isoformat()
                        elif isinstance(value, Decimal):
                            processed_row[key] = str(value)
                        elif value is None:
                            processed_row[key] = ''
                        else:
                            processed_row[key] = value
                    writer.writerow(processed_row)
            
            QMessageBox.information(self, "成功", f"已成功导出 {len(self.raw_data)} 行数据到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def _start_background_export(self, file_path: str, export_type: str):
        """启动后台导出（流式处理）"""
        if not self.original_sql or not self.main_window:
            QMessageBox.warning(self, "错误", "无法获取查询信息")
            return
        
        # 获取连接信息
        connection_id = getattr(self.main_window, 'current_connection_id', None)
        if not connection_id:
            QMessageBox.warning(self, "错误", "无法获取连接信息")
            return
        
        connection = self.main_window.db_manager.get_connection(connection_id)
        if not connection:
            QMessageBox.warning(self, "错误", "连接不存在")
            return
        
        # 移除 LIMIT 子句
        import re
        sql_no_limit = re.sub(r'\s+LIMIT\s+\d+(\s+OFFSET\s+\d+)?', '', self.original_sql, flags=re.IGNORECASE).strip().rstrip(';')
        
        # 创建进度对话框
        from PyQt6.QtWidgets import QProgressDialog
        self.export_progress = QProgressDialog(
            f"正在导出数据到 {export_type.upper()} 文件...\n已导出: 0 行",
            "取消",
            0, 0,
            self
        )
        self.export_progress.setWindowModality(Qt.WindowModality.WindowModal)
        self.export_progress.setMinimumDuration(0)
        self.export_progress.canceled.connect(self._on_export_canceled)
        self.export_progress.show()
        
        # 创建并启动导出Worker
        from src.gui.workers.export_worker import ExportWorker
        self.export_worker = ExportWorker(
            connection.get_connection_string(),
            connection.get_connect_args(),
            sql_no_limit,
            file_path,
            export_type,
            batch_size=1000  # 每批1000行
        )
        
        # 连接信号
        self.export_worker.progress_updated.connect(self._on_export_progress)
        self.export_worker.export_finished.connect(self._on_export_finished)
        
        # 启动线程
        self.export_worker.start()
    
    def _on_export_progress(self, current: int, total: int):
        """导出进度更新"""
        if hasattr(self, 'export_progress') and self.export_progress:
            if total > 0:
                self.export_progress.setMaximum(total)
                self.export_progress.setValue(current)
            self.export_progress.setLabelText(f"正在导出数据...\n已导出: {current} 行")
    
    def _on_export_finished(self, success: bool, message: str):
        """导出完成"""
        # 关闭进度对话框
        if hasattr(self, 'export_progress') and self.export_progress:
            self.export_progress.close()
            self.export_progress = None
        
        # 清理Worker
        if hasattr(self, 'export_worker') and self.export_worker:
            try:
                self.export_worker.progress_updated.disconnect()
                self.export_worker.export_finished.disconnect()
            except:
                pass
            self.export_worker.deleteLater()
            self.export_worker = None
        
        # 显示结果
        if success:
            QMessageBox.information(self, "导出成功", message)
        else:
            QMessageBox.critical(self, "导出失败", message)
    
    def _on_export_canceled(self):
        """用户取消导出"""
        if hasattr(self, 'export_worker') and self.export_worker:
            self.export_worker.stop()
    
    def _fetch_all_data(self) -> List[Dict]:
        """重新执行查询获取全部数据（不带LIMIT）"""
        if not self.original_sql or not self.main_window:
            logger.warning(f"无法获取全部数据: original_sql={bool(self.original_sql)}, main_window={bool(self.main_window)}")
            return []
        
        try:
            from sqlalchemy import create_engine, text
            
            # 获取连接信息
            connection_id = getattr(self.main_window, 'current_connection_id', None)
            if not connection_id:
                logger.warning("无法获取 connection_id")
                return []
            
            connection = self.main_window.db_manager.get_connection(connection_id)
            if not connection:
                logger.warning(f"无法获取连接: {connection_id}")
                return []
            
            # 创建引擎并执行查询
            engine = self.main_window.db_manager.get_engine(connection_id)
            if not engine:
                logger.warning(f"无法获取引擎: {connection_id}")
                return []
            
            # 移除 LIMIT 子句（使用正则表达式，处理各种情况）
            import re
            sql_no_limit = re.sub(r'\s+LIMIT\s+\d+(\s+OFFSET\s+\d+)?', '', self.original_sql, flags=re.IGNORECASE).strip().rstrip(';')
            
            logger.info(f"开始获取全部数据，原始SQL: {self.original_sql[:100]}")
            logger.info(f"移除LIMIT后的SQL: {sql_no_limit[:100]}")
            
            with engine.connect() as conn:
                result = conn.execute(text(sql_no_limit))
                # 转换为字典列表
                columns = result.keys()
                data = [dict(zip(columns, row)) for row in result.fetchall()]
                logger.info(f"成功获取 {len(data)} 行数据")
                return data
        except Exception as e:
            logger.error(f"获取全部数据失败: {str(e)}", exc_info=True)
            return []
    
    def export_to_excel(self):
        """导出为Excel"""
        if not self.raw_data:
            QMessageBox.warning(self, "警告", "没有数据可导出")
            return
        
        # 检查是否安装了 openpyxl
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(
                self,
                "缺少依赖",
                "导出Excel需要安装 openpyxl 库。\n\n请运行: pip install openpyxl"
            )
            return
        
        # 检查是否只显示了部分数据
        should_fetch_all = False
        
        # 如果自动添加了LIMIT，提示用户是否导出全部数据
        if self.auto_limit_added or self.server_side_paging:
            reply = QMessageBox.question(
                self,
                "导出选项",
                f"当前显示 {len(self.raw_data)} 行数据（已自动限制显示）。\n\n"
                f"是否要重新查询并导出全部数据？\n\n"
                f"• 是：执行完整查询，导出所有数据（后台流式处理）\n"
                f"• 否：仅导出当前显示的 {len(self.raw_data)} 行\n\n"
                f"⚠️ 提示：导出全部数据会在后台进行，不会卡住界面",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Yes
            )
            if reply == QMessageBox.StandardButton.Cancel:
                return
            should_fetch_all = (reply == QMessageBox.StandardButton.Yes)
        
        # 选择保存文件
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出为Excel",
            "",
            "Excel文件 (*.xlsx);;所有文件 (*)"
        )
        
        if not file_path:
            return
        
        # 如果只导出当前数据，使用快速方法
        if not should_fetch_all:
            self._export_current_data_to_excel(file_path)
            return
        
        # 使用后台Worker导出全部数据
        self._start_background_export(file_path, 'excel')
    
    def _export_current_data_to_excel(self, file_path: str):
        """导出当前显示的数据到Excel（同步方法）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "查询结果"
            
            columns = list(self.raw_data[0].keys())
            
            # 写入表头
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入数据
            for row_idx, row_data in enumerate(self.raw_data, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name)
                    
                    if isinstance(value, (datetime, date, time)):
                        value = value.isoformat()
                    elif isinstance(value, Decimal):
                        value = float(value)
                    elif value is None:
                        value = ''
                    
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 自动调整列宽
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            wb.save(file_path)
            
            QMessageBox.information(self, "成功", f"已成功导出 {len(self.raw_data)} 行数据到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def clear_results(self):
        """清空结果"""
        self.table.setRowCount(0)
        self.table.setColumnCount(0)
        # 状态信息显示到主窗口状态栏
        self._show_status_to_main_window("等待查询结果...")
        # 禁用导出按钮
        self.export_btn.setEnabled(False)


class MultiResultTable(QWidget):
    """多结果表格（支持Tab切换）"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
        self.result_tables: List[SingleResultTable] = []
        self.table_to_tab_index: Dict[str, int] = {}  # "connection_id:table_name" 到tab索引的映射
        self.tab_sql_map: Dict[int, str] = {}  # tab索引到SQL语句的映射
    
    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout()
        layout.setContentsMargins(5, 5, 5, 5)  # 减少外边距
        layout.setSpacing(0)  # 无间距
        self.setLayout(layout)
        
        # Tab控件
        self.tab_widget = QTabWidget()
        self.tab_widget.setTabsClosable(True)
        self.tab_widget.tabCloseRequested.connect(self.close_tab)
        # 设置tab bar的右键菜单，用于复制SQL
        self.tab_widget.tabBar().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tab_widget.tabBar().customContextMenuRequested.connect(self.on_tab_context_menu)
        # 使用事件过滤器来处理双击事件，用于复制SQL
        self.tab_widget.tabBar().installEventFilter(self)
        layout.addWidget(self.tab_widget)
    
    def _get_connection_info(self, connection_id: Optional[str]):
        """获取连接信息"""
        if not connection_id:
            return None, None
        
        main_window = getattr(self, '_main_window', None)
        if not main_window or not hasattr(main_window, 'db_manager'):
            return None, None
        
        try:
            connection = main_window.db_manager.get_connection(connection_id)
            if connection:
                return connection.get_connection_string(), connection.get_connect_args()
        except Exception as e:
            logger.error(f"获取连接信息失败: {e}")
        
        return None, None
    
    def _format_sql_title(self, sql: str, max_length: int = 40) -> str:
        """格式化SQL为tab标题"""
        sql_clean = sql.strip().replace('\n', ' ').replace('\r', ' ')
        # 移除多余空格
        sql_clean = ' '.join(sql_clean.split())
        
        if len(sql_clean) <= max_length:
            return sql_clean
        else:
            return sql_clean[:max_length] + "..."
    
    def eventFilter(self, obj, event):
        """事件过滤器，用于处理tab双击事件"""
        if obj == self.tab_widget.tabBar():
            from PyQt6.QtGui import QMouseEvent
            if event.type() == event.Type.MouseButtonDblClick:
                mouse_event = event
                if mouse_event.button() == Qt.MouseButton.LeftButton:
                    index = self.tab_widget.tabBar().tabAt(mouse_event.position().toPoint())
                    if index >= 0 and index in self.tab_sql_map:
                        self.copy_sql_to_clipboard(self.tab_sql_map[index])
                        return True
        return super().eventFilter(obj, event)
    
    def on_tab_context_menu(self, position):
        """Tab右键菜单，用于复制SQL和关闭操作"""
        tab_bar = self.tab_widget.tabBar()
        index = tab_bar.tabAt(position)
        if index < 0:
            return
        
        # 创建右键菜单
        menu = QMenu(self)
        
        # 复制SQL（如果该tab有SQL）
        if index in self.tab_sql_map:
            sql = self.tab_sql_map[index]
            copy_action = menu.addAction("复制SQL")
            copy_action.triggered.connect(lambda: self.copy_sql_to_clipboard(sql))
            menu.addSeparator()
        
        # 关闭相关菜单
        close_action = menu.addAction("关闭")
        close_action.triggered.connect(lambda: self.close_tab(index))
        
        # 如果只有一个tab，禁用"关闭其他"
        if self.tab_widget.count() > 1:
            close_others_action = menu.addAction("关闭其他")
            close_others_action.triggered.connect(lambda: self.close_other_tabs(index))
        
        # 如果只有一个tab，禁用"全部关闭"
        if self.tab_widget.count() > 1:
            menu.addSeparator()
            close_all_action = menu.addAction("全部关闭")
            close_all_action.triggered.connect(self.close_all_tabs)
        
        menu.exec(tab_bar.mapToGlobal(position))
    
    def copy_sql_to_clipboard(self, sql: str):
        """复制SQL到剪贴板"""
        from PyQt6.QtWidgets import QApplication
        from src.utils.toast_manager import show_success
        clipboard = QApplication.clipboard()
        clipboard.setText(sql)
        # 显示简短提示（Toast）
        show_success(f"✅ SQL已复制到剪贴板")
    
    def _extract_table_name(self, sql: str) -> Optional[str]:
        """
        从SQL中提取表名
        
        Args:
            sql: SQL语句
            
        Returns:
            表名，如果无法提取则返回None
        """
        sql_upper = sql.strip().upper()
        
        # 只处理SELECT查询
        if not sql_upper.startswith("SELECT"):
            return None
        
        # 尝试匹配 FROM table_name 或 FROM database.table_name
        # 匹配模式：FROM `database`.`table` 或 FROM database.table 或 FROM table
        patterns = [
            r'FROM\s+`?(\w+)`?\.`?(\w+)`?',  # FROM database.table 或 FROM `database`.`table`
            r'FROM\s+`?(\w+)`?',  # FROM table 或 FROM `table`
        ]
        
        for pattern in patterns:
            match = re.search(pattern, sql_upper, re.IGNORECASE)
            if match:
                if len(match.groups()) == 2:
                    # 返回 database.table 格式
                    return f"{match.group(1)}.{match.group(2)}"
                else:
                    # 返回表名
                    return match.group(1)
        
        return None
    
    def add_result(self, sql: str, data: Optional[List[Dict]] = None, 
                   error: Optional[str] = None, affected_rows: Optional[int] = None,
                   columns: Optional[List[str]] = None, connection_id: Optional[str] = None,
                   auto_limit_added: bool = False):
        """
        添加查询结果
        
        Args:
            sql: SQL语句（用于Tab标题）
            data: 查询结果数据
            error: 错误信息
            affected_rows: 影响的行数
            columns: 列名列表
            connection_id: 连接ID，用于区分不同连接的相同表名
        """
        # 尝试提取表名
        table_name = self._extract_table_name(sql)
        
        # 构建tab标识：使用 "connection_id:table_name" 格式，如果没有连接ID则只使用表名
        if table_name and connection_id:
            tab_key = f"{connection_id}:{table_name}"
        elif table_name:
            tab_key = table_name
        else:
            tab_key = None
        
        # 如果提取到表名，检查是否已存在该连接和表的tab
        if tab_key and tab_key in self.table_to_tab_index:
            tab_index = self.table_to_tab_index[tab_key]
            # 检查tab索引是否仍然有效
            if 0 <= tab_index < len(self.result_tables):
                # 更新现有tab的内容
                result_table = self.result_tables[tab_index]
                # 更新SQL和主窗口引用
                result_table.original_sql = sql
                result_table.main_window = getattr(self, '_main_window', None)
                result_table.auto_limit_added = auto_limit_added  # 传递自动添加LIMIT标志
                
                # 如果父级MultiResultTable有自定义的执行查询函数，传递给SingleResultTable
                if hasattr(self, '_execute_query_func'):
                    result_table.execute_query_func = self._execute_query_func
                
                # 获取连接信息
                connection_string, connect_args = self._get_connection_info(connection_id)
                
                result_table.display_results(data, error, affected_rows, columns,
                                            connection_string, connect_args)
                
                # 更新tab标题（可能SQL有变化）
                tab_title = self._format_sql_title(sql)
                self.tab_widget.setTabText(tab_index, tab_title)
                # 更新tooltip和SQL映射
                full_sql = sql.strip()
                self.tab_widget.setTabToolTip(tab_index, f"双击复制SQL\n\n{full_sql}")
                self.tab_sql_map[tab_index] = full_sql
                
                # 切换到该tab
                self.tab_widget.setCurrentIndex(tab_index)
                return
        
        # 创建新的结果表格（传递主窗口引用和SQL）
        result_table = SingleResultTable(
            parent=self,
            main_window=getattr(self, '_main_window', None),
            sql=sql
        )
        
        # 设置自动添加LIMIT标志
        result_table.auto_limit_added = auto_limit_added
        
        # 如果父级MultiResultTable有自定义的执行查询函数，传递给SingleResultTable
        if hasattr(self, '_execute_query_func'):
            result_table.execute_query_func = self._execute_query_func
        
        # 获取连接信息
        connection_string, connect_args = self._get_connection_info(connection_id)
        
        result_table.display_results(data, error, affected_rows, columns,
                                    connection_string, connect_args)
        
        # 生成Tab标题
        tab_title = self._format_sql_title(sql)
        
        # 添加Tab
        tab_index = self.tab_widget.addTab(result_table, tab_title)
        self.result_tables.append(result_table)
        
        # 设置tooltip显示完整SQL，并提示双击复制
        full_sql = sql.strip()
        self.tab_widget.setTabToolTip(tab_index, f"双击复制SQL\n\n{full_sql}")
        self.tab_sql_map[tab_index] = full_sql
        
        # 如果提取到表名，记录映射关系（使用连接ID和表名的组合）
        if tab_key:
            self.table_to_tab_index[tab_key] = tab_index
        
        # 切换到新Tab
        self.tab_widget.setCurrentIndex(tab_index)
    
    def close_tab(self, index: int):
        """关闭Tab"""
        if index < len(self.result_tables):
            # 从映射中移除该tab对应的表名
            table_name_to_remove = None
            for table_name, tab_idx in self.table_to_tab_index.items():
                if tab_idx == index:
                    table_name_to_remove = table_name
                    break
            
            if table_name_to_remove:
                del self.table_to_tab_index[table_name_to_remove]
                # 更新后续tab的索引
                for table_name in list(self.table_to_tab_index.keys()):
                    if self.table_to_tab_index[table_name] > index:
                        self.table_to_tab_index[table_name] -= 1
            
            # 从SQL映射中移除
            if index in self.tab_sql_map:
                del self.tab_sql_map[index]
                # 更新后续tab的索引
                for tab_idx in list(self.tab_sql_map.keys()):
                    if tab_idx > index:
                        self.tab_sql_map[tab_idx - 1] = self.tab_sql_map.pop(tab_idx)
            
            self.tab_widget.removeTab(index)
            self.result_tables.pop(index)
    
    def close_other_tabs(self, keep_index: int):
        """关闭除指定索引外的所有tab"""
        if keep_index < 0 or keep_index >= self.tab_widget.count():
            return
        
        # 从后往前关闭，避免索引变化
        for i in range(self.tab_widget.count() - 1, -1, -1):
            if i != keep_index:
                self.close_tab(i)
    
    def close_all_tabs(self):
        """关闭所有tab"""
        # 从后往前关闭所有tab
        for i in range(self.tab_widget.count() - 1, -1, -1):
            self.close_tab(i)
    
    def clear_all(self):
        """清空所有结果"""
        self.tab_widget.clear()
        self.result_tables.clear()
        self.table_to_tab_index.clear()
        self.tab_sql_map.clear()
    
    def display_results(
        self, 
        data: List[Dict], 
        error: Optional[str] = None,
        affected_rows: Optional[int] = None,
        sql: Optional[str] = None,
        columns: Optional[List[str]] = None
    ):
        """
        显示查询结果（兼容单结果模式）
        
        Args:
            data: 查询结果数据
            error: 错误信息
            affected_rows: 影响的行数
            sql: SQL语句（可选）
            columns: 列名列表（可选，用于无数据时显示表头）
        """
        if sql is None:
            sql = "查询结果"
        
        self.add_result(sql, data, error, affected_rows, columns)
    
    def clear_results(self):
        """清空结果（兼容旧接口）"""
        self.clear_all()
    
    def show_loading(self):
        """显示加载状态（应用于当前活动的标签页）"""
        # MultiResultTable 不需要显示加载动画
        # 因为它是多标签结构，新结果会在新标签中显示
        # 加载动画已经在 SQL 编辑器的状态栏中显示
        pass
    
    def hide_loading(self):
        """隐藏加载状态（应用于当前活动的标签页）"""
        # MultiResultTable 不需要隐藏加载动画
        pass


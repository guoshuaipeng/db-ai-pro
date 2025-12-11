"""
数据导出工作线程（支持分页查询和流式写入）
"""
from PyQt6.QtCore import QThread, pyqtSignal
from sqlalchemy import create_engine, text
from typing import List, Dict, Optional
import logging
import csv
from datetime import datetime, date, time
from decimal import Decimal

logger = logging.getLogger(__name__)


class ExportWorker(QThread):
    """数据导出工作线程"""
    
    # 定义信号
    progress_updated = pyqtSignal(int, int)  # 当前进度, 总数
    export_finished = pyqtSignal(bool, str)  # 成功/失败, 消息
    
    def __init__(self, connection_string: str, connect_args: dict, sql: str, 
                 file_path: str, export_type: str = 'csv', batch_size: int = 1000):
        super().__init__()
        self.connection_string = connection_string
        self.connect_args = connect_args
        self.sql = sql
        self.file_path = file_path
        self.export_type = export_type  # 'csv' 或 'excel'
        self.batch_size = batch_size  # 每批查询的行数
        self._should_stop = False
    
    def stop(self):
        """停止导出"""
        self._should_stop = True
        self.requestInterruption()
    
    def run(self):
        """执行导出（在工作线程中运行）"""
        engine = None
        try:
            if self.export_type == 'csv':
                self._export_to_csv()
            elif self.export_type == 'excel':
                self._export_to_excel()
        except Exception as e:
            logger.error(f"导出失败: {str(e)}", exc_info=True)
            self.export_finished.emit(False, f"导出失败: {str(e)}")
        finally:
            self.quit()
    
    def _export_to_csv(self):
        """导出为CSV（流式写入）"""
        engine = None
        try:
            # 创建数据库引擎
            engine = create_engine(
                self.connection_string,
                connect_args=self.connect_args,
                pool_pre_ping=True,
                echo=False
            )
            
            if self.isInterruptionRequested() or self._should_stop:
                return
            
            # 打开文件准备写入
            with open(self.file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = None
                total_exported = 0
                
                # 使用流式查询（yield_per）
                with engine.connect() as conn:
                    # 执行查询
                    result = conn.execution_options(stream_results=True).execute(text(self.sql))
                    
                    # 获取列名
                    columns = list(result.keys())
                    
                    # 创建CSV写入器
                    writer = csv.DictWriter(f, fieldnames=columns)
                    writer.writeheader()
                    
                    # 分批读取和写入数据
                    batch = []
                    for row in result:
                        if self.isInterruptionRequested() or self._should_stop:
                            self.export_finished.emit(False, "导出已取消")
                            return
                        
                        # 转换为字典并处理特殊类型
                        row_dict = dict(zip(columns, row))
                        processed_row = {}
                        for key, value in row_dict.items():
                            if isinstance(value, (datetime, date, time)):
                                processed_row[key] = value.isoformat()
                            elif isinstance(value, Decimal):
                                processed_row[key] = str(value)
                            elif value is None:
                                processed_row[key] = ''
                            else:
                                processed_row[key] = value
                        
                        batch.append(processed_row)
                        
                        # 每批写入
                        if len(batch) >= self.batch_size:
                            writer.writerows(batch)
                            total_exported += len(batch)
                            self.progress_updated.emit(total_exported, 0)  # 0表示未知总数
                            batch = []
                    
                    # 写入剩余数据
                    if batch:
                        writer.writerows(batch)
                        total_exported += len(batch)
                        self.progress_updated.emit(total_exported, total_exported)
            
            logger.info(f"CSV导出完成: {total_exported} 行数据")
            self.export_finished.emit(True, f"成功导出 {total_exported} 行数据到:\n{self.file_path}")
            
        except Exception as e:
            logger.error(f"CSV导出失败: {str(e)}", exc_info=True)
            self.export_finished.emit(False, f"导出失败: {str(e)}")
        finally:
            if engine:
                try:
                    engine.dispose()
                except:
                    pass
    
    def _export_to_excel(self):
        """导出为Excel（流式写入）"""
        engine = None
        try:
            # 检查是否安装了 openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
                from openpyxl.utils import get_column_letter
            except ImportError:
                self.export_finished.emit(False, "导出Excel需要安装 openpyxl 库")
                return
            
            # 创建数据库引擎
            engine = create_engine(
                self.connection_string,
                connect_args=self.connect_args,
                pool_pre_ping=True,
                echo=False
            )
            
            if self.isInterruptionRequested() or self._should_stop:
                return
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "查询结果"
            
            total_exported = 0
            
            # 使用流式查询
            with engine.connect() as conn:
                result = conn.execution_options(stream_results=True).execute(text(self.sql))
                
                # 获取列名
                columns = list(result.keys())
                
                # 写入表头
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 分批读取和写入数据
                row_idx = 2
                batch_count = 0
                for row in result:
                    if self.isInterruptionRequested() or self._should_stop:
                        self.export_finished.emit(False, "导出已取消")
                        return
                    
                    # 转换为字典并写入Excel
                    row_dict = dict(zip(columns, row))
                    for col_idx, col_name in enumerate(columns, start=1):
                        value = row_dict.get(col_name)
                        
                        # 处理特殊类型
                        if isinstance(value, (datetime, date, time)):
                            value = value.isoformat()
                        elif isinstance(value, Decimal):
                            value = float(value)
                        elif value is None:
                            value = ''
                        
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    row_idx += 1
                    total_exported += 1
                    batch_count += 1
                    
                    # 每批更新进度
                    if batch_count >= self.batch_size:
                        self.progress_updated.emit(total_exported, 0)
                        batch_count = 0
            
            # 自动调整列宽（限制最大宽度）
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # 保存文件
            wb.save(self.file_path)
            
            logger.info(f"Excel导出完成: {total_exported} 行数据")
            self.export_finished.emit(True, f"成功导出 {total_exported} 行数据到:\n{self.file_path}")
            
        except Exception as e:
            logger.error(f"Excel导出失败: {str(e)}", exc_info=True)
            self.export_finished.emit(False, f"导出失败: {str(e)}")
        finally:
            if engine:
                try:
                    engine.dispose()
                except:
                    pass

